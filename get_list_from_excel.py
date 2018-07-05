import openpyxl
import os
import tempfile
import zipfile
import re


def fix_xlsx(in_file):
    '''
    приводит xlsx файл из 1С к стандартной структуре
    '/xl/sharedStrings.xml' -> '/xl/SharedStrings.xml'
    :param in_file:
    :return:
    '''
    tmpfd, tmp = tempfile.mkstemp(dir=os.path.dirname(in_file))
    os.close(tmpfd)
    _filename = '[Content_Types].xml'
    data = ''
    with zipfile.ZipFile(in_file, 'r') as zin:
        with zipfile.ZipFile(tmp, 'w') as zout:
            for item in zin.infolist():
                if item.filename != _filename:
                    zout.writestr(item, zin.read(item.filename))
                else:
                    data = zin.read(_filename).decode()
    os.remove(in_file)
    os.rename(tmp, in_file)
    data = data.replace('/xl/sharedStrings.xml', '/xl/SharedStrings.xml')
    with zipfile.ZipFile(in_file, mode='a', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(_filename, data)



class Item:
    def __init__(self):
        self.name = ''
        self.left = 0.0
        self.code = ''
        self.art = ''
        self.price_opt = 0.0
        self.ed_opt = ''
        self.price_ros = 0.0
        self.ed_ros = ''
        self.photo_path = ''

class Header:
    def __init__(self):
        self.name = ''




class XlsxItems:
    def __init__(self, filename, first_row=14):
        self.__filename = filename
        self.__list_items = []
        self.__first_row = first_row


    def __prepare_image_folder(self):
        pass


    def get_items(self):
        '''
        генерация списка словарей на основании XLSX файла
        :return
        '''
        try:
            wb = openpyxl.load_workbook(filename=self.__filename)
        except KeyError:
            fix_xlsx(filename)
            wb = openpyxl.load_workbook(filename=self.__filename)
        except Exception as ex:
            print(ex)

        sheet = wb.active
        row = self.__first_row
        while True:
            item = Item()
            cell = sheet.cell(row=row, column=2)
            if cell.value is None:
                break
            if sheet.cell(row=row, column=5).value is None:
                header = Header()
                header.name = cell.value
                self.__list_items.append(header)
                row += 1
                continue

            item.name = re.sub(r'\s+', ' ', sheet.cell(row=row, column=2).value).lstrip()
            item.art = sheet.cell(row=row, column=5).value
            item.code = sheet.cell(row=row, column=4).value.replace(' ', '')
            item.ed_opt = sheet.cell(row=row, column=8).value
            item.ed_ros = sheet.cell(row=row, column=10).value
            item.left = sheet.cell(row=row, column=3).value
            item.price_opt = sheet.cell(row=row, column=7).value
            item.price_ros = sheet.cell(row=row, column=9).value

            row += 1
            self.__list_items.append(item)
        return self.__list_items








if __name__ == '__main__':
    filename = 'prise list 07,06.2018.xlsx'
    x = XlsxItems(filename)
    listx = x.get_items()
    for row in listx:
        if isinstance(row, Item):
            print(' {} {} {} {} {} {} {} {}'.format(row.name, row.left,  row.code, row.art, row.price_opt, row.ed_opt, row.price_ros, row.ed_ros))
        #if isinstance(row, Header):
        #    print('{}'.format(row.name))
    print('Total in XLSX file: {}'.format(len(listx)))
